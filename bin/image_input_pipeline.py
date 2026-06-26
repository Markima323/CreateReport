from __future__ import annotations

import argparse
import base64
import io
import json
import re
import sys
import tempfile
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

import pymupdf
from openpyxl import load_workbook
from PIL import Image, ImageOps, UnidentifiedImageError

from process_excel_to_word import (
    DEFAULT_GEMINI_MODEL,
    clean_text,
    create_gemini_client,
    get_interaction_output_text,
    load_gemini_api_key,
    load_rules,
)

for stream in (sys.stdout, sys.stderr):
    if hasattr(stream, "reconfigure"):
        stream.reconfigure(encoding="utf-8", errors="replace")


DIRECT_IMAGE_MIME_TYPES = {
    ".jpeg": "image/jpeg",
    ".jpg": "image/jpeg",
    ".jfif": "image/jpeg",
    ".png": "image/png",
    ".webp": "image/webp",
    ".heic": "image/heic",
    ".heif": "image/heif",
}
CONVERTED_IMAGE_SUFFIXES = {
    ".avif",
    ".bmp",
    ".dib",
    ".gif",
    ".ico",
    ".pbm",
    ".pcx",
    ".pgm",
    ".ppm",
    ".tga",
    ".tif",
    ".tiff",
}
PDF_SUFFIXES = {".pdf"}
SUPPORTED_IMAGE_SUFFIXES = (
    set(DIRECT_IMAGE_MIME_TYPES) | CONVERTED_IMAGE_SUFFIXES | PDF_SUFFIXES
)
MAX_IMAGE_FRAMES_PER_FILE = 50
PDF_RENDER_DPI = 200
SUPPORTED_EXCEL_SUFFIXES = {".xlsx", ".xlsm"}

RECORD_FIELDS = [
    "序号",
    "报告号",
    "支行",
    "姓名",
    "基准日本金",
    "基准日利息",
    "基准日本息",
    "地址",
    "面积",
    "第一顺位",
    "登记类型",
    "性质",
    "权利价值",
    "预计回收",
    "评估单价",
    "评估价值",
    "快速变现系数",
    "快速变现价值",
    "清收完成时间",
    "折现期限",
    "折现系数",
    "折现价值",
    "处置费用",
    "安置费用",
    "可回收金额",
    "债权评估值",
    "偿债率",
    "权证编号",
    "证载权利人",
    "保证人",
]

DEBTOR_FIELDS = [
    "姓名",
    "性别",
    "民族",
    "身份证号码",
    "住址",
]

OPTIONAL_GUARANTOR_FIELDS = [
    "保证人统一社会信用代码",
    "保证人类型",
    "保证人法定代表人",
    "保证人成立日期",
    "保证人营业场所",
    "保证人经营范围",
]

EXCEL_RECORD_FIELDS = [
    *RECORD_FIELDS,
    *[
        "债务人" + str(index) + ("" if field == "姓名" else field)
        for index in range(1, 5)
        for field in DEBTOR_FIELDS
    ],
    *OPTIONAL_GUARANTOR_FIELDS,
]

GENERATION_REQUIRED_FIELDS = [
    "序号",
    "支行",
    "姓名",
    "基准日本金",
    "基准日利息",
    "地址",
    "面积",
    "权利价值",
    "预计回收",
    "评估单价",
    "快速变现系数",
    "折现期限",
    "处置费用",
    "安置费用",
    "权证编号",
    "证载权利人",
    "债务人1",
    "债务人1性别",
    "债务人1民族",
    "债务人1身份证号码",
    "债务人1住址",
]

IMAGE_EXTRACTION_SCHEMA: Dict[str, Any] = {
    "type": "object",
    "properties": {
        "地址": {"type": "string"},
        "面积": {"type": "string"},
        "第一顺位": {"type": "string"},
        "登记类型": {"type": "string"},
        "权利价值": {"type": "string"},
        "权证编号": {"type": "string"},
        "证载权利人": {"type": "string"},
        "保证人": {"type": "string"},
        "债务人": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "姓名": {"type": "string"},
                    "性别": {"type": "string"},
                    "民族": {"type": "string"},
                    "身份证号码": {"type": "string"},
                    "住址": {"type": "string"},
                },
                "required": DEBTOR_FIELDS,
            },
        },
        "warnings": {
            "type": "array",
            "items": {"type": "string"},
        },
    },
    "required": [
        "地址",
        "面积",
        "第一顺位",
        "登记类型",
        "权利价值",
        "权证编号",
        "证载权利人",
        "保证人",
        "债务人",
        "warnings",
    ],
}


def parse_folder_name(folder_name: str) -> Tuple[str, str, str]:
    match = re.fullmatch(r"\s*(\d+)-(.+?)-(.+?)\s*", folder_name)
    if not match:
        raise ValueError(
            f"文件夹名称格式错误: {folder_name}，应为“序号-支行-姓名”"
        )
    sequence, branch, name = (part.strip() for part in match.groups())
    if branch.endswith("支行"):
        branch = branch[:-2].strip()
    if not branch or not name:
        raise ValueError(f"文件夹名称缺少支行或姓名: {folder_name}")
    return sequence, branch, name


def list_person_folders(input_dir: Path) -> List[Path]:
    if not input_dir.is_dir():
        raise FileNotFoundError(f"图片输入目录不存在: {input_dir}")
    folders = sorted(
        (
            path
            for path in input_dir.iterdir()
            if path.is_dir()
            and re.fullmatch(r"\s*\d+-.+?-.+?\s*", path.name)
            and any(
                item.is_file()
                and item.suffix.lower() in SUPPORTED_IMAGE_SUFFIXES
                for item in path.iterdir()
            )
        ),
        key=lambda path: (
            int(parse_folder_name(path.name)[0]),
            path.name,
        ),
    )
    if not folders:
        raise ValueError(
            "输入根目录中没有类型 1 人员子文件夹；"
            "文件夹应按“序号-支行-姓名”命名并包含图片: "
            f"{input_dir}"
        )
    return folders


def list_images(folder: Path) -> List[Path]:
    images = sorted(
        path
        for path in folder.iterdir()
        if path.is_file() and path.suffix.lower() in SUPPORTED_IMAGE_SUFFIXES
    )
    if len(images) < 2:
        ignored_files = sorted(
            path.name
            for path in folder.iterdir()
            if path.is_file()
            and not path.name.startswith("~$")
            and path.suffix.lower() not in SUPPORTED_IMAGE_SUFFIXES
        )
        ignored_note = (
            f"；未识别文件: {', '.join(ignored_files)}"
            if ignored_files
            else ""
        )
        supported = ", ".join(sorted(SUPPORTED_IMAGE_SUFFIXES))
        raise ValueError(
            f"{folder.name} 至少需要两张图片，当前识别到 {len(images)} 张"
            f"{ignored_note}。支持格式: {supported}"
        )
    return images


def find_input_workbook(
    report_dir: Path,
    fallback_dir: Optional[Path] = None,
) -> Path:
    workbooks = sorted(
        path
        for path in report_dir.iterdir()
        if (
            path.is_file()
            and not path.name.startswith("~$")
            and path.suffix.lower() in SUPPORTED_EXCEL_SUFFIXES
        )
    )
    if not workbooks and fallback_dir is not None:
        fallback_dir = fallback_dir.resolve()
        if fallback_dir != report_dir.resolve():
            return find_input_workbook(fallback_dir)
    if not workbooks:
        raise FileNotFoundError(
            f"报告数据文件夹中未找到 Excel 文件（.xlsx/.xlsm）: {report_dir}"
        )
    if len(workbooks) > 1:
        raise ValueError(
            f"报告数据文件夹中存在多个 Excel 文件，请只保留一个（{report_dir}）: "
            + ", ".join(path.name for path in workbooks)
        )
    return workbooks[0]


def normalize_excel_header(value: Any) -> str:
    return re.sub(r"\s+", "", clean_text(value))


def normalize_person_name(value: Any) -> str:
    return re.sub(r"\s+", "", clean_text(value))


def normalize_branch(value: Any) -> str:
    branch = normalize_person_name(value)
    return branch[:-2] if branch.endswith("支行") else branch


def normalize_sequence(value: Any) -> str:
    text = clean_text(value)
    if re.fullmatch(r"\d+\.0+", text):
        return text.split(".", 1)[0]
    return text


def normalize_excel_value(value: Any) -> Any:
    if isinstance(value, (date, datetime)):
        return clean_text(value)
    return normalize_scalar(value)


def load_excel_records(workbook_file: Path) -> List[Dict[str, Any]]:
    workbook = load_workbook(
        filename=workbook_file,
        read_only=True,
        data_only=True,
    )
    records: List[Dict[str, Any]] = []
    expected_headers = {
        normalize_excel_header(field): field
        for field in EXCEL_RECORD_FIELDS
    }
    try:
        for worksheet in workbook.worksheets:
            header_row_number = 0
            header_columns: Dict[str, int] = {}
            for row_number, row in enumerate(
                worksheet.iter_rows(min_row=1, max_row=10, values_only=True),
                start=1,
            ):
                candidate_columns = {
                    normalize_excel_header(value): index
                    for index, value in enumerate(row)
                    if normalize_excel_header(value)
                }
                if (
                    "姓名" in candidate_columns
                    and "序号" in candidate_columns
                    and "支行" in candidate_columns
                ):
                    header_row_number = row_number
                    header_columns = candidate_columns
                    break
            if not header_row_number:
                continue

            for row_number, row in enumerate(
                worksheet.iter_rows(
                    min_row=header_row_number + 1,
                    values_only=True,
                ),
                start=header_row_number + 1,
            ):
                data: Dict[str, Any] = {}
                for normalized_header, field in expected_headers.items():
                    column_index = header_columns.get(normalized_header)
                    if column_index is None or column_index >= len(row):
                        continue
                    data[field] = normalize_excel_value(row[column_index])
                if not normalize_person_name(data.get("姓名")):
                    continue
                records.append(
                    {
                        "file": str(workbook_file),
                        "sheet": worksheet.title,
                        "row": row_number,
                        "data": data,
                    }
                )
    finally:
        workbook.close()
    if not records:
        raise ValueError(
            f"Excel 中未找到包含“序号、支行、姓名”表头的数据表: {workbook_file}"
        )
    return records


def match_excel_record(
    records: Sequence[Dict[str, Any]],
    sequence: str,
    branch: str,
    name: str,
) -> Dict[str, Any]:
    expected_name = normalize_person_name(name)
    expected_sequence = normalize_sequence(sequence)
    expected_branch = normalize_branch(branch)
    candidates = [
        item
        for item in records
        if normalize_person_name(item["data"].get("姓名")) == expected_name
    ]
    if not candidates:
        raise ValueError(
            f"Excel 中找不到姓名为“{name}”的数据行"
        )

    if len(candidates) > 1:
        sequence_matches = [
            item
            for item in candidates
            if normalize_sequence(item["data"].get("序号")) == expected_sequence
        ]
        if sequence_matches:
            candidates = sequence_matches
    if len(candidates) > 1:
        branch_matches = [
            item
            for item in candidates
            if normalize_branch(item["data"].get("支行")) == expected_branch
        ]
        if branch_matches:
            candidates = branch_matches
    if len(candidates) != 1:
        locations = ", ".join(
            f"{item['sheet']}!{item['row']}" for item in candidates
        )
        raise ValueError(
            f"Excel 中“{name}”匹配到多行，无法唯一确定: {locations}"
        )

    matched = candidates[0]
    data = matched["data"]
    actual_sequence = normalize_sequence(data.get("序号"))
    actual_branch = normalize_branch(data.get("支行"))
    if actual_sequence and expected_sequence and actual_sequence != expected_sequence:
        raise ValueError(
            f"Excel 姓名“{name}”的序号为 {actual_sequence}，"
            f"与文件夹序号 {expected_sequence} 不一致"
        )
    if actual_branch and expected_branch and actual_branch != expected_branch:
        raise ValueError(
            f"Excel 姓名“{name}”的支行为“{clean_text(data.get('支行'))}”，"
            f"与文件夹支行“{branch}”不一致"
        )
    return matched


def normalize_image_frame(frame: Image.Image) -> Image.Image:
    normalized = ImageOps.exif_transpose(frame)
    if normalized.mode == "P":
        return normalized.convert(
            "RGBA" if "transparency" in normalized.info else "RGB"
        )
    if normalized.mode in {"1", "L", "LA", "RGB", "RGBA"}:
        return normalized
    return normalized.convert("RGB")


def prepare_pdf_parts(path: Path) -> List[Tuple[str, bytes, str]]:
    try:
        document = pymupdf.open(path)
    except Exception as exc:
        raise ValueError(f"无法读取 PDF 文件: {path}，{exc}") from exc
    try:
        if document.needs_pass:
            raise ValueError(f"PDF 已加密，需要先移除密码: {path}")
        page_count = document.page_count
        if page_count < 1:
            raise ValueError(f"PDF 没有可识别页面: {path}")
        if page_count > MAX_IMAGE_FRAMES_PER_FILE:
            raise ValueError(
                f"PDF 页数过多: {path.name} 共 {page_count} 页，"
                f"最多支持 {MAX_IMAGE_FRAMES_PER_FILE} 页"
            )
        parts: List[Tuple[str, bytes, str]] = []
        for page_index in range(page_count):
            page = document.load_page(page_index)
            pixmap = page.get_pixmap(dpi=PDF_RENDER_DPI, alpha=False)
            parts.append(
                (
                    f"{path.name}（第{page_index + 1}页）",
                    pixmap.tobytes("png"),
                    "image/png",
                )
            )
        return parts
    finally:
        document.close()


def prepare_image_parts(
    path: Path,
) -> List[Tuple[str, bytes, str]]:
    suffix = path.suffix.lower()
    direct_mime_type = DIRECT_IMAGE_MIME_TYPES.get(suffix)
    if direct_mime_type:
        return [(path.name, path.read_bytes(), direct_mime_type)]
    if suffix in PDF_SUFFIXES:
        return prepare_pdf_parts(path)
    if suffix not in CONVERTED_IMAGE_SUFFIXES:
        raise ValueError(f"不支持的图片格式: {path}")

    try:
        with Image.open(path) as image:
            frame_count = int(getattr(image, "n_frames", 1) or 1)
            if frame_count > MAX_IMAGE_FRAMES_PER_FILE:
                raise ValueError(
                    f"图片页数过多: {path.name} 共 {frame_count} 页，"
                    f"最多支持 {MAX_IMAGE_FRAMES_PER_FILE} 页"
                )
            parts: List[Tuple[str, bytes, str]] = []
            for frame_index in range(frame_count):
                image.seek(frame_index)
                frame = normalize_image_frame(image.copy())
                buffer = io.BytesIO()
                frame.save(buffer, format="PNG", optimize=True)
                label = (
                    path.name
                    if frame_count == 1
                    else f"{path.name}（第{frame_index + 1}页）"
                )
                parts.append((label, buffer.getvalue(), "image/png"))
            return parts
    except UnidentifiedImageError as exc:
        raise ValueError(f"无法读取图片文件: {path}") from exc
    except OSError as exc:
        raise ValueError(f"转换图片失败: {path}，{exc}") from exc


def build_extraction_prompt(
    sequence: str,
    branch: str,
    name: str,
) -> str:
    return (
        "你是中国不动产登记簿和居民身份证的高精度文字提取器。"
        "请联合读取本次请求中的全部图片；图片可能旋转、倒置或包含身份证正反面，"
        "应先识别正确方向。只提取图片中明确可见的信息，禁止根据常识、文件夹名称"
        "或其他字段编造图片未出现的内容。\n\n"
        f"文件夹元数据：序号={sequence}，支行={branch}，姓名={name}。"
        "元数据只用于核对人员归属，不得覆盖图片中的身份证或登记簿文字。\n\n"
        "字段规则：\n"
        "1. 地址取登记簿中房屋坐落/坐落的完整文字。\n"
        "2. 面积取房屋建筑面积，输出平方米数值，不带单位。\n"
        "3. 权证编号取登记簿顶部“权证号”对应的不动产权证/证明编号；"
        "不要误取抵押登记证明号。\n"
        "4. 证载权利人取登记簿权利人；多人用中文顿号连接。\n"
        "5. 权利价值取工商银行对应抵押登记的债权数额，原图为万元时换算为元，"
        "只输出数值。\n"
        "6. 第一顺位仅在工商银行明确为首位/第一条有效抵押权时输出“是”；"
        "无法确认时输出空字符串。\n"
        "7. 登记类型根据登记簿判断为“预抵押登记”或“正式抵押登记”；"
        "无法确认时输出空字符串。\n"
        "8. 债务人按身份证图片逐人输出姓名、性别、民族、身份证号码、住址。"
        "民族不追加“族”。身份证号码保持原字符，末位 X 使用大写。\n"
        "9. 保证人仅在图片明确出现保证人名称时填写，否则返回空字符串。\n"
        "10. 无法确认的字段必须返回空字符串，并在 warnings 中说明。"
    )


def build_image_input(
    image_paths: Sequence[Path],
    prompt: str,
) -> List[Dict[str, Any]]:
    items: List[Dict[str, Any]] = [{"type": "text", "text": prompt}]
    for image_path in image_paths:
        for label, image_bytes, mime_type in prepare_image_parts(image_path):
            items.append(
                {
                    "type": "text",
                    "text": f"图片文件名：{label}",
                }
            )
            items.append(
                {
                    "type": "image",
                    "data": base64.b64encode(image_bytes).decode("ascii"),
                    "mime_type": mime_type,
                    "resolution": "high",
                }
            )
    return items


def extract_folder_images(
    client: Any,
    model: str,
    thinking_level: str,
    folder: Path,
    store_interactions: bool,
) -> Dict[str, Any]:
    sequence, branch, name = parse_folder_name(folder.name)
    images = list_images(folder)
    interaction = client.create(
        model=model,
        input=build_image_input(
            images,
            build_extraction_prompt(sequence, branch, name),
        ),
        response_format={
            "type": "text",
            "mime_type": "application/json",
            "schema": IMAGE_EXTRACTION_SCHEMA,
        },
        generation_config={"thinking_level": thinking_level},
        store=store_interactions,
    )
    if interaction.get("status") == "failed":
        raise RuntimeError(
            f"Gemini interaction 失败: {interaction.get('error') or interaction}"
        )
    output_text = get_interaction_output_text(interaction)
    if not output_text:
        raise RuntimeError("Gemini 返回了空响应")
    try:
        extracted = json.loads(output_text)
    except json.JSONDecodeError as exc:
        raise RuntimeError(f"Gemini 返回的结构化 JSON 无法解析: {exc}") from exc
    return {
        "sequence": sequence,
        "branch": branch,
        "name": name,
        "folder": folder.name,
        "images": [path.name for path in images],
        "interaction_id": clean_text(interaction.get("id")),
        "model": clean_text(interaction.get("model")) or model,
        "usage": (
            interaction.get("usage")
            if isinstance(interaction.get("usage"), dict)
            else {}
        ),
        "extracted": extracted,
    }


def normalize_scalar(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip().replace(",", "")
    if not text:
        return ""
    if re.fullmatch(r"-?\d+", text):
        return int(text)
    if re.fullmatch(r"-?\d+\.\d+", text):
        return float(text)
    return str(value).strip()


def load_supplements(path: Path) -> Dict[str, Dict[str, Any]]:
    if not path.is_file():
        return {}
    payload = json.loads(path.read_text(encoding="utf-8-sig"))
    if not isinstance(payload, dict):
        raise ValueError(f"图片补充数据必须是 JSON 对象: {path}")
    return payload


def build_record(
    extraction: Dict[str, Any],
    supplement: Dict[str, Any],
    excel_match: Dict[str, Any],
) -> Dict[str, Any]:
    extracted = extraction["extracted"]
    record: Dict[str, Any] = {field: "" for field in RECORD_FIELDS}
    record.update(
        {
            "序号": extraction["sequence"],
            "支行": extraction["branch"],
            "姓名": extraction["name"],
        }
    )
    for field in (
        "地址",
        "面积",
        "第一顺位",
        "登记类型",
        "权利价值",
        "权证编号",
        "证载权利人",
        "保证人",
    ):
        record[field] = normalize_scalar(extracted.get(field))

    for index in range(1, 5):
        for field in DEBTOR_FIELDS:
            key = "债务人" + str(index) + ("" if field == "姓名" else field)
            record[key] = ""
    debtors = extracted.get("债务人") or []
    for index, debtor in enumerate(debtors[:4], start=1):
        if not isinstance(debtor, dict):
            continue
        record[f"债务人{index}"] = clean_text(debtor.get("姓名"))
        for field in ("性别", "民族", "身份证号码", "住址"):
            record[f"债务人{index}{field}"] = clean_text(debtor.get(field))

    for field, value in supplement.items():
        if field.startswith("_"):
            continue
        if field not in record or record[field] in ("", None):
            record[field] = value

    excel_imported_fields: List[str] = []
    excel_overridden_fields: List[Dict[str, Any]] = []
    for field, value in excel_match["data"].items():
        normalized_value = normalize_excel_value(value)
        if normalized_value in ("", None):
            continue
        previous_value = record.get(field)
        if previous_value not in ("", None):
            if clean_text(previous_value) != clean_text(normalized_value):
                excel_overridden_fields.append(
                    {
                        "field": field,
                        "previous": previous_value,
                        "excel": normalized_value,
                    }
                )
        else:
            excel_imported_fields.append(field)
        record[field] = normalized_value

    missing_fields = [
        field
        for field in GENERATION_REQUIRED_FIELDS
        if record.get(field) in ("", None)
    ]
    warnings = [
        clean_text(item)
        for item in extracted.get("warnings", [])
        if clean_text(item)
    ]
    return {
        "folder": extraction["folder"],
        "images": extraction["images"],
        "gemini": {
            "interaction_id": extraction["interaction_id"],
            "model": extraction["model"],
            "usage": extraction["usage"],
        },
        "excel": {
            "file": excel_match["file"],
            "sheet": excel_match["sheet"],
            "row": excel_match["row"],
            "imported_fields": excel_imported_fields,
            "overridden_fields": excel_overridden_fields,
        },
        "data": record,
        "missing_fields": missing_fields,
        "warnings": warnings,
        "status": "ready" if not missing_fields else "incomplete",
    }


def write_json_atomic(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(
        mode="w",
        encoding="utf-8",
        suffix=".json",
        dir=path.parent,
        delete=False,
    ) as handle:
        temp_path = Path(handle.name)
        json.dump(payload, handle, ensure_ascii=False, indent=2)
        handle.write("\n")
    try:
        temp_path.replace(path)
    finally:
        if temp_path.exists():
            temp_path.unlink()


def friendly_extraction_error(exc: Exception) -> str:
    error_text = str(exc)
    lowered = error_text.lower()
    if "prepayment credits are depleted" in lowered:
        return (
            "Gemini API 429（too_many_requests / RESOURCE_EXHAUSTED）；"
            "API 详细原因：预付费额度已耗尽。请在 Google AI Studio 充值，"
            "或更换有可用额度的 API Key"
        )
    if "gemini api http 429" in lowered:
        return (
            "Gemini API 429（too_many_requests / RESOURCE_EXHAUSTED）；"
            "可能超过 RPM、TPM、RPD 或消费额度限制，请稍后重试并检查 "
            "Google AI Studio 的项目限额"
        )
    if "api key" in lowered and any(
        marker in lowered
        for marker in ("invalid", "expired", "not valid")
    ):
        return "Gemini API Key 无效或已失效，请更换后重试"
    return error_text


def extract_input_directory(
    input_dir: Path,
    output_file: Path,
    individual_dir: Path,
    supplements_file: Path,
    api_key: str,
    model: str,
    thinking_level: str,
    store_interactions: bool,
    only_folder: Optional[Path] = None,
) -> Dict[str, Any]:
    client = create_gemini_client(api_key)
    supplements = load_supplements(supplements_file)
    records: List[Dict[str, Any]] = []
    errors: List[Dict[str, Any]] = []
    if only_folder is not None:
        folder = only_folder.resolve()
        if not folder.is_dir():
            raise FileNotFoundError(f"人员图片文件夹不存在: {folder}")
        parse_folder_name(folder.name)
        folders = [folder]
    else:
        folders = list_person_folders(input_dir)
    print(f"发现人员文件夹: {len(folders)}")
    workbook_files: List[str] = []
    for index, folder in enumerate(folders, start=1):
        print(f"[{index}/{len(folders)}] 提取图片: {folder.name}")
        try:
            workbook_file = find_input_workbook(folder, fallback_dir=input_dir)
            workbook_files.append(str(workbook_file))
            excel_records = load_excel_records(workbook_file)
            print(f"  Excel 数据源: {workbook_file}")
            extraction = extract_folder_images(
                client=client,
                model=model,
                thinking_level=thinking_level,
                folder=folder,
                store_interactions=store_interactions,
            )
            print(f"  图片识别完成，开始匹配 Excel: {workbook_file.name}")
            excel_match = match_excel_record(
                excel_records,
                extraction["sequence"],
                extraction["branch"],
                extraction["name"],
            )
            print(
                f"  Excel 匹配: {excel_match['sheet']}!"
                f"{excel_match['row']}（Excel 非空字段优先）"
            )
            record = build_record(
                extraction,
                supplements.get(folder.name, {}),
                excel_match,
            )
            records.append(record)
            write_json_atomic(
                individual_dir / f"{folder.name}.json",
                record,
            )
            print(
                f"  JSON: {record['status']} | "
                f"缺失字段 {len(record['missing_fields'])}"
            )
        except Exception as exc:
            display_error = friendly_extraction_error(exc)
            errors.append({"folder": folder.name, "error": display_error})
            print(f"  ERROR: {display_error}")
            error_text = str(exc).lower()
            if any(
                marker in error_text
                for marker in (
                    "prepayment credits are depleted",
                    "insufficient credits",
                    "billing account is not active",
                )
            ):
                for skipped_folder in folders[index:]:
                    errors.append(
                        {
                            "folder": skipped_folder.name,
                            "error": "因 Gemini 计费/余额错误，未继续处理本批次",
                        }
                    )
                print("  已停止后续图片提取，请检查 Gemini 项目余额或更换 API Key")
                break

    result = {
        "schema_version": "2.0",
        "source_type": "report_folders_with_images_and_excel",
        "source_root": str(input_dir),
        "excel_files": sorted(set(workbook_files)),
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "model": model,
        "records": records,
        "errors": errors,
    }
    write_json_atomic(output_file, result)
    print(f"合并 JSON: {output_file}")
    return result


def build_parser(project_root: Path) -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Extract report records from person image folders using Gemini."
    )
    parser.add_argument(
        "--input-dir",
        default=str(project_root / "Input" / "1"),
    )
    parser.add_argument(
        "--folder",
        default=None,
        help="仅处理一个人员文件夹，用于串行生成流程。",
    )
    parser.add_argument(
        "--output-file",
        default=str(project_root / "bin" / "json" / "图片提取数据.json"),
    )
    parser.add_argument(
        "--individual-dir",
        default=str(project_root / "bin" / "json" / "人员数据"),
    )
    parser.add_argument(
        "--supplements-file",
        default=str(
            project_root
            / "bin"
            / "template"
            / "1"
            / "图片输入补充数据.json"
        ),
    )
    parser.add_argument(
        "--rules-file",
        default=str(
            project_root
            / "bin"
            / "template"
            / "1"
            / "价值分析报告生成规则.json"
        ),
    )
    parser.add_argument("--model", default=None)
    parser.add_argument(
        "--api-key-file",
        default=str(project_root / "gemini_api.txt"),
    )
    return parser


def main(argv: Optional[Sequence[str]] = None) -> int:
    project_root = Path(__file__).resolve().parent.parent
    args = build_parser(project_root).parse_args(argv)
    input_dir = Path(args.input_dir).expanduser().resolve()
    output_file = Path(args.output_file).expanduser().resolve()
    individual_dir = Path(args.individual_dir).expanduser().resolve()
    supplements_file = Path(args.supplements_file).expanduser().resolve()
    rules_file = Path(args.rules_file).expanduser().resolve()
    api_key_file = Path(args.api_key_file).expanduser().resolve()
    only_folder = (
        Path(args.folder).expanduser().resolve()
        if args.folder
        else None
    )

    rules = load_rules(rules_file)
    ai_config = rules.get("ai", {})
    model = clean_text(args.model) or clean_text(ai_config.get("model")) or DEFAULT_GEMINI_MODEL
    thinking_level = clean_text(ai_config.get("thinking_level")) or "medium"
    store_interactions = bool(ai_config.get("store_interactions", False))
    api_key = load_gemini_api_key(api_key_file)

    result = extract_input_directory(
        input_dir=input_dir,
        output_file=output_file,
        individual_dir=individual_dir,
        supplements_file=supplements_file,
        api_key=api_key,
        model=model,
        thinking_level=thinking_level,
        store_interactions=store_interactions,
        only_folder=only_folder,
    )
    incomplete = sum(
        1 for record in result["records"] if record["status"] != "ready"
    )
    print(f"提取成功: {len(result['records'])}")
    print(f"不完整记录: {incomplete}")
    print(f"错误: {len(result['errors'])}")
    return 0 if not result["errors"] and not incomplete else 1


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        raise SystemExit(1)
